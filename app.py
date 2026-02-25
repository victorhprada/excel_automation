import pandas as pd
import openpyxl
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border
from openpyxl.cell.cell import MergedCell
from copy import copy
from datetime import date
from dateutil.relativedelta import relativedelta
import re
from datetime import datetime
from openpyxl.utils import column_index_from_string
from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import io
from fastapi import HTTPException

# ========================================
# Funções Auxiliares
# ========================================

app = FastAPI(title="API de Validação de Faturamento Excel")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
async def root():
    return {
        "status": "online",
        "mensagem": "🚀 API de Validação de Faturamento rodando perfeitamente!",
        "versao": "1.0"
    }

@app.post("/api/processar-comissao")
async def processar_comissao(
    arquivo_base: UploadFile = File(...),
    arquivo_parceiro: UploadFile = File(...),
    
    # 2. Recebendo os Textos (Mês e Datas)
    target_month: str = Form(...),
    data_inicio: str = Form(...),
    data_fim: str = Form(...)
):
    try:
        # =================================================================
        # ETAPA 1: LER ARQUIVOS PARA A MEMÓRIA RAM
        # =================================================================
        base_bytes = io.BytesIO(await arquivo_base.read())
        parceiro_bytes = io.BytesIO(await arquivo_parceiro.read())
        print("📥 1. Arquivos recebidos pelo Render! Lendo para a memória...")
        
        parceiro_wb = openpyxl.load_workbook(parceiro_bytes, data_only=True)
        base_wb = openpyxl.load_workbook(base_bytes, data_only=False)
        print("⚙️ 2. Carregando Workbooks no OpenPyxl (Isso pode demorar)...")

        # Converter datas recebidas do Front-end (String) para objetos Date do Python
        # Assumindo que o Lovable vai mandar no formato HTML padrão "YYYY-MM-DD"
        dt_inicio = datetime.strptime(data_inicio, "%Y-%m-%d").date()
        dt_fim = datetime.strptime(data_fim, "%Y-%m-%d").date()
        
        print("🚀 3. Iniciando limpeza e regras de negócio...")
        # =================================================================
        # ETAPA 2: LÓGICA DE NEGÓCIO PURA (Sem comandos Streamlit)
        # =================================================================
        
        # 2.1 Validar Abas
        valido, mensagem = validar_abas_necessarias(parceiro_wb, base_wb)
        if not valido:
            raise HTTPException(status_code=400, detail=f"Erro de validação: {mensagem}")
            
        template_existe, msg_template = validar_template_jan26(base_wb)
        if not template_existe:
            raise HTTPException(status_code=400, detail=f"Template ausente: {msg_template}")

        # 2.2 Clonar Template JAN.26
        if target_month in base_wb.sheetnames:
            del base_wb[target_month]
            
        ws_mes = base_wb.copy_worksheet(base_wb['JAN.26'])
        ws_mes.title = target_month

        # 2.3 Limpar e Inserir Dados do Parceiro
        limpar_dados_worksheet(ws_mes, manter_linha_1=True)
        inserir_dados_colunas_especificas(parceiro_wb['Parcelas Pagas'], ws_mes, 1, 13, 2)
        aplicar_regras_colunas_n_x(ws_mes, target_month, 2)

        # 2.4 Atualizar Aba BASE
        ultima_linha_base = encontrar_ultima_linha(base_wb['BASE'])
        linha_inicio_append = ultima_linha_base + 1
        
        copiar_producao_para_base(parceiro_wb['Produção'], base_wb['BASE'])
        atualizar_aba_base(base_wb, parceiro_wb, target_month, linha_inicio_append)

        # 2.6 Recarregar BASE no Pandas para pegar novos registros
        ws_base_ativa = base_wb['BASE']
        data = list(ws_base_ativa.values)
        if data:
            cols = data[0]
            base_df_atualizado = pd.DataFrame(data[1:], columns=cols)
        else:
            base_df_atualizado = pd.DataFrame()

        # 2.7 Ciclo de Validação e Inserção de Novos Inadimplentes
        # A processar_ciclo_validacao também precisa ter os st. removidos por dentro!
        processar_ciclo_validacao(
            base_df_atualizado, base_wb, target_month, dt_inicio, dt_fim
        )

        # 2.8 Atualizar Aba RESUMO
        if 'RESUMO' in base_wb.sheetnames:
            coluna_alvo = atualizar_resumo_mes_faturamento(base_wb, target_month)
            atualizar_resumo_ciclo_pmt(base_wb, target_month)
            verificar_e_corrigir_headers_regras(base_wb['RESUMO'])
            atualizar_resumo_bloco_final(base_wb, target_month, col_idx=coluna_alvo)

        # =================================================================
        # ETAPA 3: DEVOLVER O ARQUIVO PRONTO PARA O LOVABLE
        # =================================================================
        output = io.BytesIO()
        base_wb.save(output)
        output.seek(0)
        
        nome_arquivo = f"Processado_{target_month}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={nome_arquivo}"}
        )

    except Exception as e:
        # Captura qualquer erro de código e devolve para o front-end exibir um Toast vermelho
        raise HTTPException(status_code=500, detail=f"Erro interno no servidor: {str(e)}")

def copiar_estilo(celula_origem, celula_destino):
    """
    Copia atributos de formatação de uma célula para outra.
    
    Atributos copiados: font, border, fill, number_format, alignment
    
    Reconstrução manual de Border para evitar RecursionError com StyleProxy.
    
    Args:
        celula_origem: Célula de onde copiar o estilo
        celula_destino: Célula para onde copiar o estilo
    """
    if celula_origem.has_style:
        celula_destino.font = copy(celula_origem.font)
        
        # Cópia manual segura para evitar RecursionError em StyleProxy
        b_origem = celula_origem.border
        if b_origem:
            celula_destino.border = Border(
                left=copy(b_origem.left),
                right=copy(b_origem.right),
                top=copy(b_origem.top),
                bottom=copy(b_origem.bottom),
                diagonal=copy(b_origem.diagonal),
                diagonal_direction=b_origem.diagonal_direction,
                outline=b_origem.outline,
                vertical=b_origem.vertical,
                horizontal=b_origem.horizontal
            )
        
        celula_destino.fill = copy(celula_origem.fill)
        celula_destino.number_format = celula_origem.number_format
        celula_destino.alignment = copy(celula_origem.alignment)


def encontrar_coluna_por_header(ws, nome_header):
    """
    Busca dinamicamente o índice de uma coluna pelo nome do cabeçalho (linha 1).
    
    Args:
        ws: Worksheet onde buscar
        nome_header: Nome exato do cabeçalho a procurar (case-sensitive)
    
    Returns:
        int: Índice da coluna (1-based) ou None se não encontrar
    """
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header == nome_header:
            return col
    
    # Se não encontrar, retornar None (permitir ao chamador decidir)
    return None


def validar_abas_necessarias(parceiro_wb, base_wb):
    """
    Valida se todas as abas necessárias existem nos workbooks.
    Retorna (sucesso: bool, mensagem: str)
    """
    abas_parceiro_necessarias = ['Parcelas Pagas', 'Produção']
    abas_base_necessarias = ['BASE', 'INADIMPLENTES', 'JAN.26']  # Incluir JAN.26 como template
    
    # Verificar PARCEIRO
    for aba in abas_parceiro_necessarias:
        if aba not in parceiro_wb.sheetnames:
            return False, f"Aba '{aba}' não encontrada no arquivo PARCEIRO"
    
    # Verificar BASE
    for aba in abas_base_necessarias:
        if aba not in base_wb.sheetnames:
            return False, f"Aba '{aba}' não encontrada no arquivo BASE"
    
    return True, "Todas as abas necessárias estão presentes (incluindo template JAN.26)"


def encontrar_ultima_linha(ws):
    """
    Encontra a última linha preenchida em uma worksheet.
    Retorna o número da linha.
    """
    for row in range(ws.max_row, 0, -1):
        # Verificar se há algum valor não-nulo na linha
        if any(ws.cell(row=row, column=col).value is not None 
               for col in range(1, ws.max_column + 1)):
            return row
    return 0  # Se worksheet vazia, retornar 0


def calcular_mes_anterior(mes_str):
    """
    Calcula o mês anterior a partir de target_month (ex: 'JAN.26').

    Converte para datetime (dia 1), subtrai 1 mês e formata como 'mmm/yy'
    em português (ex: 'dez/25'). JAN.26 -> dez/25.

    Args:
        mes_str: String no formato 'MMM.AA' (ex: 'JAN.26')

    Returns:
        str: Mês anterior no formato 'mmm/yy' (ex: 'dez/25')
    """
    meses_eng = {
        'JAN': 1, 'FEV': 2, 'MAR': 3, 'ABR': 4, 'MAI': 5, 'JUN': 6,
        'JUL': 7, 'AGO': 8, 'SET': 9, 'OUT': 10, 'NOV': 11, 'DEZ': 12
    }
    meses_pt = {
        1: 'jan', 2: 'fev', 3: 'mar', 4: 'abr', 5: 'mai', 6: 'jun',
        7: 'jul', 8: 'ago', 9: 'set', 10: 'out', 11: 'nov', 12: 'dez'
    }
    partes = mes_str.upper().strip().split('.')
    mes_abrev = partes[0]
    ano_2d = int(partes[1])
    ano = 2000 + ano_2d
    mes_num = meses_eng[mes_abrev]
    d = date(ano, mes_num, 1)
    if mes_num == 1:
        d_ant = date(ano - 1, 12, 1)
    else:
        d_ant = date(ano, mes_num - 1, 1)
    return f"{meses_pt[d_ant.month]}/{str(d_ant.year)[-2:]}"


def encontrar_ultima_coluna_resumo(ws):
    """
    Encontra o índice da última coluna preenchida na linha 2 da aba RESUMO.

    Usado para determinar onde inserir a nova coluna de Mês Faturamento.

    Args:
        ws: Worksheet (aba RESUMO)

    Returns:
        int: Índice 1-based da última coluna com valor na linha 2, ou 1 se vazia.
    """
    ultima = 1
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=2, column=col).value is not None:
            ultima = col
    return ultima


def atualizar_resumo_mes_faturamento(base_wb, target_month):
    """
    Atualiza o bloco MÊS FATURAMENTO (linhas 2 a 6) na aba RESUMO.

    Insere uma nova coluna à direita da última preenchida na linha 2,
    preenche valores e fórmulas (SUMIF/COUNTIF na BASE, comissão 3%),
    e copia o estilo da coluna anterior.

    Args:
        base_wb: Workbook do arquivo BASE (deve conter aba 'RESUMO')
        target_month: String do mês (ex: 'JAN.26')

    Returns:
        int: Índice (1-based) da coluna criada/reutilizada
    """
    ws_resumo = base_wb['RESUMO']
    ultima_col = encontrar_ultima_coluna_resumo(ws_resumo)
    nova_coluna = ultima_col + 1
    
    # Verificação inteligente: se a coluna já está vazia, reutilizar (evita gap)
    linha2_vazia = ws_resumo.cell(row=2, column=nova_coluna).value is None
    linha9_vazia = ws_resumo.cell(row=9, column=nova_coluna).value is None
    linha9_valor = ws_resumo.cell(row=9, column=nova_coluna).value
    
    # Se ambas vazias E não for o header 'REGRA PARA PARCELAMENTO', reutilizar coluna
    eh_header_regras = linha9_valor and 'REGRA' in str(linha9_valor).upper()
    
    if not (linha2_vazia and linha9_vazia and not eh_header_regras):
        # Coluna tem dados ou é header importante: inserir nova coluna
        ws_resumo.insert_cols(nova_coluna)
    
    letra = get_column_letter(nova_coluna)

    mes_faturado = target_month.replace('.', '/').lower()
    mes_ref = calcular_mes_anterior(target_month)

    ws_resumo.cell(row=2, column=nova_coluna, value=mes_faturado)
    ws_resumo.cell(row=3, column=nova_coluna, value=mes_ref)
    ws_resumo.cell(row=4, column=nova_coluna, value=f"=SUMIF(BASE!$K:$K,RESUMO!{letra}3,BASE!$D:$D)")
    ws_resumo.cell(row=5, column=nova_coluna, value=f"=COUNTIF(BASE!$K:$K,RESUMO!{letra}3)")
    ws_resumo.cell(row=6, column=nova_coluna, value=f"={letra}4*3%")

    # Busca inteligente da coluna molde (ignora colunas vazias intermediárias)
    col_molde = nova_coluna - 1
    while col_molde >= 1:
        if ws_resumo.cell(row=4, column=col_molde).value is not None:
            break
        col_molde -= 1
    
    # Copiar estilo da coluna molde (se encontrada)
    if col_molde >= 1:
        for r in range(2, 7):
            celula_origem = ws_resumo.cell(row=r, column=col_molde)
            celula_destino = ws_resumo.cell(row=r, column=nova_coluna)
            copiar_estilo(celula_origem, celula_destino)
    
    return nova_coluna


def atualizar_resumo_ciclo_pmt(base_wb, target_month):
    """
    Atualiza o bloco CICLO PMT (linhas 9 a 18) na aba RESUMO.

    Reutiliza a coluna criada pelo bloco Mês Faturamento (linha 2).
    Calcula período de 4 meses antes do target_month (dia 23 ao dia 20),
    preenche fórmulas COUNTIFS/SUMIFS na BASE e na aba do mês,
    e copia formatação da coluna anterior.

    Args:
        base_wb: Workbook do arquivo BASE (deve conter aba 'RESUMO')
        target_month: String do mês (ex: 'JAN.26')

    Returns:
        None
    """
    ws_resumo = base_wb['RESUMO']
    
    # Formatar target_month para o padrão da linha 2: 'jan/26'
    mes_faturado = target_month.replace('.', '/').lower()
    
    # Localizar coluna pelo cabeçalho da linha 2
    col_idx = None
    for col in range(1, ws_resumo.max_column + 1):
        valor_celula = ws_resumo.cell(row=2, column=col).value
        if valor_celula and str(valor_celula).strip().lower() == mes_faturado:
            col_idx = col
            break
    
    if not col_idx:
        raise ValueError(f"Coluna com '{mes_faturado}' não encontrada na linha 2 da aba RESUMO")
    
    # Converter target_month para date e subtrair 4 meses
    meses_eng = {
        'JAN': 1, 'FEV': 2, 'MAR': 3, 'ABR': 4, 'MAI': 5, 'JUN': 6,
        'JUL': 7, 'AGO': 8, 'SET': 9, 'OUT': 10, 'NOV': 11, 'DEZ': 12
    }
    partes = target_month.upper().strip().split('.')
    mes_num = meses_eng[partes[0]]
    ano = 2000 + int(partes[1])
    
    data_ref = date(ano, mes_num, 1)
    data_ref_menos_4 = data_ref - relativedelta(months=4)
    
    # Datas do ciclo: dia 23 (início) e dia 20 do mês seguinte (fim)
    data_ini = date(data_ref_menos_4.year, data_ref_menos_4.month, 23)
    data_fim_mes = data_ref_menos_4 + relativedelta(months=1)
    data_fim = date(data_fim_mes.year, data_fim_mes.month, 20)
    
    # Strings formatadas para fórmulas Excel
    data_ini_str = data_ini.strftime("%d/%m/%Y")
    data_fim_str = data_fim.strftime("%d/%m/%Y")
    
    # Header: '23/09 a 20/10 - 2025'
    header_str = f"{data_ini.strftime('%d/%m')} a {data_fim.strftime('%d/%m')} - {data_ini.year}"
    
    letra = get_column_letter(col_idx)
    
    # Preencher linhas 9 a 18 na coluna alinhada
    ws_resumo.cell(row=9, column=col_idx, value=header_str)
    ws_resumo.cell(row=10, column=col_idx, value=f'=COUNTIFS(BASE!$H:$H,">={data_ini_str}",BASE!$H:$H,"<={data_fim_str}")')
    ws_resumo.cell(row=11, column=col_idx, value=f'=SUMIFS(BASE!$D:$D,BASE!$H:$H,">={data_ini_str}",BASE!$H:$H,"<={data_fim_str}")')
    ws_resumo.cell(row=12, column=col_idx, value=f"=SUM('{target_month}'!L:L)")
    ws_resumo.cell(row=13, column=col_idx, value=f"=COUNTA('{target_month}'!O:O)-1")
    ws_resumo.cell(row=14, column=col_idx, value=f'=COUNTIFS(\'{target_month}\'!R:R,">={data_ini_str}",\'{target_month}\'!R:R,"<={data_fim_str}")')
    ws_resumo.cell(row=15, column=col_idx, value=f"={letra}13-{letra}14")
    ws_resumo.cell(row=16, column=col_idx, value=None)  # Vazio
    ws_resumo.cell(row=17, column=col_idx, value=f"={letra}14-{letra}10")
    
    # Linha 18: copiar fórmula da célula esquerda se houver
    celula_esq_18 = ws_resumo.cell(row=18, column=col_idx - 1)
    if celula_esq_18.value:
        ws_resumo.cell(row=18, column=col_idx, value=celula_esq_18.value)
    else:
        ws_resumo.cell(row=18, column=col_idx, value=None)
    
    # Busca inteligente da coluna molde (ignora colunas vazias intermediárias)
    col_molde = col_idx - 1
    while col_molde >= 1:
        if ws_resumo.cell(row=10, column=col_molde).value is not None:
            break
        col_molde -= 1
    
    # Copiar estilo da coluna molde (se encontrada)
    if col_molde >= 1:
        for r in range(9, 19):
            celula_origem = ws_resumo.cell(row=r, column=col_molde)
            celula_destino = ws_resumo.cell(row=r, column=col_idx)
            copiar_estilo(celula_origem, celula_destino)


def verificar_e_corrigir_headers_regras(ws):
    """
    Restaura os cabeçalhos da tabela REGRA PARA PARCELAMENTO que podem sumir
    após inserções de colunas.
    
    Procura 'REGRA PARA PARCELAMENTO' na linha 9 e força os valores dos headers
    nas colunas seguintes com formatação de cabeçalho.
    
    Args:
        ws: Worksheet da aba RESUMO
    
    Returns:
        None
    """
    # Procurar 'REGRA PARA PARCELAMENTO' na linha 9
    col_regra = None
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=9, column=col).value
        if valor and 'REGRA' in str(valor).upper() and 'PARCELAMENTO' in str(valor).upper():
            col_regra = col
            break
    
    if not col_regra:
        return  # Se não encontrar, não faz nada
    
    # Forçar valores dos headers nas colunas seguintes
    headers = [
        'CICLO PARCELAS',
        'Repasse DataPrev p/Paketa',
        'Receita Wiipo'
    ]
    
    for i, header in enumerate(headers, start=1):
        col_atual = col_regra + i
        
        # CRÍTICO: Remover mesclagem e limpar cache ANTES de escrever
        coord = f"{get_column_letter(col_atual)}9"
        
        # Verificar se essa coordenada está em alguma mesclagem
        for merged_range in list(ws.merged_cells.ranges):
            if coord in merged_range:
                ws.unmerge_cells(str(merged_range))
                print(f"✅ DEBUG: Mesclagem {merged_range} removida para liberar header {coord}")
                
                # Deletar o cache da célula para forçar recriação
                if (9, col_atual) in ws._cells:
                    del ws._cells[(9, col_atual)]
                    print(f"🔄 DEBUG: Cache da célula {coord} limpo")
                
                break
        
        # Agora sim escrever o valor
        celula = ws.cell(row=9, column=col_atual)
        celula.value = header
        
        # Aplicar estilo de cabeçalho (copiar da coluna REGRA PARA PARCELAMENTO)
        celula_origem = ws.cell(row=9, column=col_regra)
        copiar_estilo(celula_origem, celula)


def preparar_celula_para_escrita(ws, row, col):
    """
    Verifica se a célula alvo é uma MergedCell (parte de uma mesclagem).
    Se for, identifica o intervalo pai e DESFAZ (unmerge) para liberar a escrita.
    
    Args:
        ws: Worksheet onde verificar
        row: Linha da célula (1-based)
        col: Coluna da célula (1-based)
    
    Returns:
        None
    """
    cell = ws.cell(row=row, column=col)
    
    # Verifica se a célula está em algum intervalo mesclado
    for merged_range in list(ws.merged_cells.ranges):
        if cell.coordinate in merged_range:
            ws.unmerge_cells(str(merged_range))
            print(f"DEBUG: Mesclagem {merged_range} removida para liberar a célula {cell.coordinate}")
            break


def atualizar_resumo_bloco_final(base_wb, target_month, col_idx):
    """
    Atualiza o bloco FATURAMENTO (linhas 20 a 23).
    Estratégia: Ler da linha 2 + Destravar linha 20 + Escrever.
    
    CRÍTICO: Remove mesclagens ANTES de acessar qualquer célula para evitar 
    erro 'MergedCell' object attribute 'value' is read-only'.
    
    Imita o processo manual:
    1. Lê o valor da linha 2 (já preenchida por atualizar_resumo_mes_faturamento)
    2. Destrava células usando coordenadas string (sem acessar objetos célula)
    3. Escreve o valor lido + fórmulas
    
    Args:
        base_wb: Workbook do arquivo BASE (deve conter aba 'RESUMO')
        target_month: String do mês (ex: 'JAN.26') - usado apenas para referência
        col_idx: Índice (1-based) da coluna onde escrever os dados
    
    Returns:
        None
    """
    ws = base_wb['RESUMO']
    letra = get_column_letter(col_idx)
    
    print(f"DEBUG: Iniciando Bloco Final na Coluna {col_idx} ({letra})")
    
    # PASSO A: Ler valor da linha 2 (já preenchida anteriormente)
    valor_linha2 = ws.cell(row=2, column=col_idx).value
    
    if not valor_linha2:
        print(f"⚠️ AVISO: Linha 2 da coluna {letra} está vazia!")
        # Fallback: usar target_month formatado
        valor_linha2 = target_month.replace('.', '/').lower()
    
    print(f"DEBUG: Valor lido da linha 2: '{valor_linha2}'")
    
    # PASSO B: CRÍTICO - Remover mesclagens SEM acessar células
    linhas_alvo = [20, 21, 22, 23]
    
    for linha_num in linhas_alvo:
        # Construir coordenada como string (ex: "L20") SEM chamar ws.cell()
        coord = f"{letra}{linha_num}"
        
        # Verificar se essa coordenada está em alguma mesclagem
        for merged_range in list(ws.merged_cells.ranges):
            if coord in merged_range:
                ws.unmerge_cells(str(merged_range))
                print(f"✅ DEBUG: Mesclagem {merged_range} removida para liberar {coord}")
                
                # CRÍTICO: Deletar o cache da célula para forçar recriação
                if (linha_num, col_idx) in ws._cells:
                    del ws._cells[(linha_num, col_idx)]
                    print(f"🔄 DEBUG: Cache da célula {coord} limpo")
                
                break
    
    # PASSO C: Escrever dados
    try:
        # L20: Colar o valor lido da linha 2
        ws.cell(row=20, column=col_idx).value = valor_linha2
        
        # L21: Referência ao topo (Comissão Originação) -> ={LETRA}6
        ws.cell(row=21, column=col_idx).value = f"={letra}6"
        
        # L22: Referência ao meio (Comissão Parcelas) -> ={LETRA}12
        ws.cell(row=22, column=col_idx).value = f"={letra}12"
        
        # L23: Soma -> =SUM({LETRA}21:{LETRA}22)
        ws.cell(row=23, column=col_idx).value = f"=SUM({letra}21:{letra}22)"
        
        print(f"✅ DEBUG: Dados escritos com sucesso na coluna {letra}")
    except Exception as e:
        print(f"❌ ERRO CRÍTICO NA ESCRITA: {e}")
        raise
    
    # PASSO D: Clonar estilo da coluna anterior (Format Painter)
    try:
        # Busca inteligente da coluna molde (ignora colunas vazias intermediárias)
        col_anterior = col_idx - 1
        while col_anterior >= 1:
            # Verifica se a linha 20 (header) tem valor (indicador de coluna preenchida)
            if ws.cell(row=20, column=col_anterior).value is not None:
                break
            col_anterior -= 1
        
        if col_anterior >= 1:
            # Copiar largura da coluna
            letra_anterior = get_column_letter(col_anterior)
            ws.column_dimensions[letra].width = ws.column_dimensions[letra_anterior].width
            print(f"📏 DEBUG: Largura da coluna {letra} copiada de {letra_anterior}")
            
            # Copiar estilo de cada célula (linhas 20-23)
            for r in linhas_alvo:
                source = ws.cell(row=r, column=col_anterior)
                target = ws.cell(row=r, column=col_idx)
                if source.has_style:
                    try:
                        copiar_estilo(source, target)
                    except:
                        pass
    except Exception as e:
        print(f"⚠️ Erro ao copiar estilo: {e}")


def copiar_dados_aba(ws_origem, ws_destino, incluir_header=False):
    """
    Copia todos os dados de uma worksheet origem para destino.
    Usa encontrar_ultima_linha() para escrever na posição correta,
    evitando problemas com formatação em células vazias.
    """
    linhas_copiadas = 0
    start_row_origem = 1 if incluir_header else 2  # Pular header se incluir_header=False
    
    # Encontrar a próxima linha vazia no destino
    ultima_linha_destino = encontrar_ultima_linha(ws_destino)
    proxima_linha_destino = ultima_linha_destino + 1
    
    # Se destino estiver completamente vazio, começar da linha 1
    if ultima_linha_destino == 0:
        proxima_linha_destino = 1
    
    # Iterar sobre linhas da origem
    for row in ws_origem.iter_rows(min_row=start_row_origem, values_only=True):
        # Pular linhas completamente vazias
        if all(cell is None for cell in row):
            continue
        
        # Escrever célula por célula na linha de destino
        for col_idx, valor in enumerate(row, start=1):
            ws_destino.cell(row=proxima_linha_destino, column=col_idx, value=valor)
        
        proxima_linha_destino += 1
        linhas_copiadas += 1
    
    return linhas_copiadas


def copiar_producao_para_base(ws_origem, ws_destino):
    """
    Copia dados da aba 'Produção' para 'BASE' de forma explícita e controlada.
    
    CRÍTICO: Usa mapeamento segmentado de colunas:
    - A-G (1-7): Cópia direta origem -> destino
    - H (8): Fórmula injetada =F{row} (não vem da origem)
    - H-J origem (8-10) -> I-K destino (9-11): Deslocamento +1
    
    Não copia formatação ou fórmulas da origem (exceto fórmula injetada em H).
    Copia a formatação da última linha existente na BASE para manter consistência visual.
    
    Args:
        ws_origem: Worksheet de origem (Produção)
        ws_destino: Worksheet de destino (BASE)
    
    Returns:
        int: Número de linhas copiadas
    """
    # 1. Encontrar última linha real em BASE (onde coluna A tem valor)
    last_row_base = 0
    for row in range(1, ws_destino.max_row + 1):
        if ws_destino.cell(row=row, column=1).value is not None:
            last_row_base = row
    
    # Se BASE está vazia, começar da linha 2 (linha 1 é header)
    if last_row_base == 0:
        last_row_base = 1
    
    new_row = last_row_base + 1
    linhas_copiadas = 0
    
    # 2. Iterar sobre linhas da aba 'Produção' (começando da linha 2)
    for source_row in range(2, ws_origem.max_row + 1):
        # Verificar se linha tem dados na coluna A (se não, parar)
        if ws_origem.cell(row=source_row, column=1).value is None:
            break
        
        # 3. Copiar colunas com mapeamento segmentado
        # Etapa 3.1: Colunas A-G (1-7) - Cópia direta
        for col in range(1, 8):  # 1 a 7 (A até G)
            valor = ws_origem.cell(row=source_row, column=col).value
            cell_nova = ws_destino.cell(row=new_row, column=col, value=valor)
            
            # Copiar formatação da linha molde
            if last_row_base > 1:
                cell_molde = ws_destino.cell(row=last_row_base, column=col)
                copiar_estilo(cell_molde, cell_nova)
        
        # Etapa 3.2: Coluna H (8) - Injetar fórmula =F{row}
        cell_nova = ws_destino.cell(row=new_row, column=8, value=f"=F{new_row}")
        
        # Copiar formatação da linha molde
        if last_row_base > 1:
            cell_molde = ws_destino.cell(row=last_row_base, column=8)
            copiar_estilo(cell_molde, cell_nova)
        
        # Etapa 3.3: Colunas H-J da origem (8-10) -> I-K do destino (9-11)
        # Deslocamento: origem_col + 1 = destino_col
        for origem_col in range(8, 11):  # 8, 9, 10 (H, I, J da origem)
            destino_col = origem_col + 1  # 9, 10, 11 (I, J, K do destino)
            valor = ws_origem.cell(row=source_row, column=origem_col).value
            cell_nova = ws_destino.cell(row=new_row, column=destino_col, value=valor)
            
            # Copiar formatação da linha molde
            if last_row_base > 1:
                cell_molde = ws_destino.cell(row=last_row_base, column=destino_col)
                copiar_estilo(cell_molde, cell_nova)
        
        new_row += 1
        linhas_copiadas += 1
    
    return linhas_copiadas


def filtrar_inadimplentes(ws_origem, coluna_validacao='VALIDAÇÃO'):
    """
    Filtra linhas onde a coluna VALIDAÇÃO é igual a 'Não'.
    Retorna lista de tuplas com os dados das linhas filtradas.
    """
    inadimplentes = []
    
    # Encontrar índice da coluna VALIDAÇÃO no header (linha 1)
    header_row = list(ws_origem.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    
    try:
        col_idx = header_row.index(coluna_validacao)
    except ValueError:
        raise ValueError(f"Coluna '{coluna_validacao}' não encontrada na aba")
    
    # Filtrar linhas onde VALIDAÇÃO = 'Não'
    for row in ws_origem.iter_rows(min_row=2, values_only=True):  # Pular header
        if row[col_idx] == 'Não':
            inadimplentes.append(row)
    
    return inadimplentes


def validar_template_jan26(workbook):
    """
    Valida se a aba 'JAN.26' (template padrão) existe no workbook BASE.
    Esta aba é usada como matriz para criar todas as novas abas de mês.
    
    Args:
        workbook: Workbook do openpyxl
    
    Returns:
        tuple: (existe: bool, mensagem: str)
    """
    template_nome = 'JAN.26'
    
    if template_nome in workbook.sheetnames:
        return True, f"Template '{template_nome}' encontrado"
    else:
        return False, f"ERRO CRÍTICO: Aba '{template_nome}' não encontrada. Esta aba é necessária como template padrão."


def capturar_formulas_colunas(ws, linha=2, col_inicio=17, col_fim=24):
    """
    Captura fórmulas de colunas específicas de uma linha.
    Retorna dicionário {coluna_idx: formula_string}
    
    Args:
        ws: Worksheet do openpyxl
        linha: Linha de onde extrair fórmulas (default: 2)
        col_inicio: Primeira coluna (default: 17 = Q)
        col_fim: Última coluna (default: 24 = X)
    
    Returns:
        dict: {col_idx: formula} apenas para colunas que têm fórmulas
    """
    formulas = {}
    
    for col_idx in range(col_inicio, col_fim + 1):
        cell = ws.cell(row=linha, column=col_idx)
        
        # Verificar se a célula tem fórmula
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            formulas[col_idx] = cell.value
    
    return formulas


def atualizar_formula_linha(formula, linha_nova):
    """
    Atualiza referências de linha em uma fórmula Excel.
    
    Args:
        formula: String da fórmula (ex: '=VLOOKUP(@Q:Q;BASE!A:K;11;0)')
        linha_nova: Número da nova linha
    
    Returns:
        str: Fórmula com referências de linha atualizadas
    
    Exemplos:
        atualizar_formula_linha('=IF(ISNUMBER(MATCH(V2;Q:Q;0));"Sim";"Não")', 5)
        -> '=IF(ISNUMBER(MATCH(V5;Q:Q;0));"Sim";"Não")'
    """
    import re
    
    # Padrão para referências de célula com linha específica (ex: A2, V2, Q2)
    # Captura letra(s) seguida(s) de número
    padrao = r'([A-Z]+)(\d+)'
    
    def substituir_linha(match):
        coluna = match.group(1)
        # Substituir qualquer número de linha pelo novo
        return f"{coluna}{linha_nova}"
    
    # Substituir todas as referências de linha na fórmula
    formula_atualizada = re.sub(padrao, substituir_linha, formula)
    
    return formula_atualizada


def limpar_dados_worksheet(ws, manter_linha_1=True):
    """
    Limpa todos os dados de uma worksheet, mantendo a linha 1 (header).
    
    Args:
        ws: Worksheet do openpyxl
        manter_linha_1: Se True, mantém linha 1 intacta
    """
    linha_inicial = 2 if manter_linha_1 else 1
    
    # Iterar de trás para frente para evitar problemas com índices
    for row_idx in range(ws.max_row, linha_inicial - 1, -1):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).value = None


def aplicar_regras_colunas_n_x(ws, target_month, linha_inicio=2):
    """
    Aplica regras de negócio explícitas para as colunas N até X.
    
    Args:
        ws: Worksheet do openpyxl onde aplicar as regras
        target_month: String do mês alvo (ex: 'JAN.26')
        linha_inicio: Linha inicial (default: 2, primeira linha de dados)
    
    Returns:
        dict: {'linhas_n_o': int, 'linhas_q_w': int, 'ccbs_unicos': int}
    
    Regras:
        MOMENTO A (Colunas N-O para todas as linhas):
            Col N (14) - Mês Faturado: target_month formatado (minúsculo, hífen)
            Col O (15) - Data Desembolso: =VLOOKUP(A{row},'BASE'!A:H,8,0)
            Col P (16) - Separador: None (vazio)
        
        MOMENTO B (Colunas Q-W apenas para CCBs únicos):
            Col Q (17) - CCB: Valor único da coluna A
            Col R (18) - Mês Originação: =VLOOKUP(Q{row},'BASE'!A:K,11,0)
            Col S (19) - Repasse: =SUMIF(A:A,Q{row},L:L)
            Col T (20) - Data Desemb 1: =VLOOKUP(Q{row},'BASE'!A:H,8,0)
            Col U (21) - Separador: None (vazio)
            Col V (22) e W (23): None (vazio)
            Col X (24) - Vazio (removida fórmula)
    """
    # ========================================
    # PREPARAÇÃO: Formatar target_month
    # ========================================
    # Converter 'JAN.26' -> 'jan-26' (minúsculo com hífen)
    mes_faturado = target_month.replace('.', '-').lower()
    
    # ========================================
    # ENCONTRAR ÚLTIMA LINHA COM DADOS
    # ========================================
    ultima_linha = linha_inicio - 1
    for row in range(linha_inicio, ws.max_row + 1):
        if ws.cell(row=row, column=1).value is not None:
            ultima_linha = row
        else:
            break
    
    if ultima_linha < linha_inicio:
        return {'linhas_n_o': 0, 'linhas_q_w': 0, 'ccbs_unicos': 0}
    
    # ========================================
    # MOMENTO A: Preencher Colunas N-O (todas as linhas)
    # ========================================
    linhas_n_o = 0
    
    for row in range(linha_inicio, ultima_linha + 1):
        # Col N (14) - Mês Faturado: String formatada
        ws.cell(row=row, column=14, value=mes_faturado)
        
        # Col O (15) - Data Desembolso: Fórmula VLOOKUP
        ws.cell(row=row, column=15, value=f"=VLOOKUP(A{row},'BASE'!A:H,8,0)")
        
        # Col P (16) - Separador: Vazio
        ws.cell(row=row, column=16, value=None)
        
        linhas_n_o += 1
    
    # ========================================
    # MOMENTO B: Preencher Colunas Q-W (apenas CCBs únicos)
    # ========================================
    
    # Extrair todos os valores da coluna A (CCBs)
    ccbs_todos = []
    for row in range(linha_inicio, ultima_linha + 1):
        valor_a = ws.cell(row=row, column=1).value
        if valor_a is not None:
            ccbs_todos.append(valor_a)
    
    # Gerar lista de CCBs únicos (preservando ordem de primeira aparição)
    ccbs_unicos = []
    vistos = set()
    for ccb in ccbs_todos:
        if ccb not in vistos:
            ccbs_unicos.append(ccb)
            vistos.add(ccb)
    
    # Preencher colunas Q-W para cada CCB único
    linhas_q_w = 0
    row_destino = linha_inicio
    
    for ccb_unico in ccbs_unicos:
        # Col Q (17) - CCB: Valor único
        ws.cell(row=row_destino, column=17, value=ccb_unico)
        
        # Col R (18) - Mês Originação: Fórmula VLOOKUP
        ws.cell(row=row_destino, column=18, value=f"=VLOOKUP(Q{row_destino},'BASE'!A:K,11,0)")
        
        # Col S (19) - Repasse: Fórmula SUMIF
        ws.cell(row=row_destino, column=19, value=f"=SUMIF(A:A,Q{row_destino},L:L)")
        
        # Col T (20) - Data Desemb 1: Fórmula VLOOKUP
        ws.cell(row=row_destino, column=20, value=f"=VLOOKUP(Q{row_destino},'BASE'!A:H,8,0)")
        
        # Col U (21) - Separador: Vazio
        ws.cell(row=row_destino, column=21, value=None)
        
        # Col V (22) e W (23): Vazios
        ws.cell(row=row_destino, column=22, value=None)
        ws.cell(row=row_destino, column=23, value=None)
        
        # Col X (24) - Vazio (sem fórmula)
        ws.cell(row=row_destino, column=24, value=None)
        
        row_destino += 1
        linhas_q_w += 1
    
    return {
        'linhas_n_o': linhas_n_o,
        'linhas_q_w': linhas_q_w,
        'ccbs_unicos': len(ccbs_unicos)
    }


def encontrar_colunas_meses(ws_base):
    """
    Identifica colunas de meses na aba BASE.
    
    Returns:
        list: [
            {'nome': 'Setembro', 'indice': 17, 'letra': 'Q'},
            {'nome': 'Outubro', 'indice': 18, 'letra': 'R'},
            ...
        ]
    """
    colunas_meses = []
    
    # Encontrar índice da coluna P (última coluna antes dos meses)
    col_p_index = 16  # P = 16
    
    # Encontrar índice da coluna DATA dinamicamente
    col_data_index = encontrar_coluna_por_header(ws_base, 'DATA')
    
    if not col_data_index:
        # Fallback: assumir que está após a última coluna
        col_data_index = ws_base.max_column + 1
        # Log de aviso (não gera erro pois esta função é só para mapear meses)
    
    # Iterar entre P+1 e DATA-1
    for col_idx in range(col_p_index + 1, col_data_index):
        header = ws_base.cell(row=1, column=col_idx).value
        if header:  # Se tem cabeçalho, é coluna de mês
            colunas_meses.append({
                'nome': header,
                'indice': col_idx,
                'letra': get_column_letter(col_idx)
            })
    
    return colunas_meses


def inserir_coluna_mes(ws_base, target_month, colunas_meses):
    """
    Insere nova coluna de mês na aba BASE.
    
    Args:
        ws_base: Worksheet da BASE
        target_month: String do mês (ex: 'JAN.26')
        colunas_meses: Lista de colunas de meses existentes
    
    Returns:
        dict: {'nome': 'JAN.26', 'indice': 22, 'letra': 'V'}
    """
    # Determinar posição de inserção
    if colunas_meses:
        # Inserir após a última coluna de mês
        ultimo_mes_idx = colunas_meses[-1]['indice']
        pos_insercao = ultimo_mes_idx + 1
    else:
        # Se não há colunas de meses, inserir após P
        pos_insercao = 17  # Q
    
    # Inserir coluna
    ws_base.insert_cols(pos_insercao)
    
    # Definir cabeçalho
    ws_base.cell(row=1, column=pos_insercao, value=target_month)
    
    # Aplicar fórmula COUNTIF em todas as linhas (da linha 2 até última)
    ultima_linha = encontrar_ultima_linha(ws_base)
    
    for row in range(2, ultima_linha + 1):
        # Fórmula: =COUNTIF('JAN.26'!A:A, BASE!A2)
        formula = f"=COUNTIF('{target_month}'!A:A,BASE!A{row})"
        ws_base.cell(row=row, column=pos_insercao, value=formula)
    
    return {
        'nome': target_month,
        'indice': pos_insercao,
        'letra': get_column_letter(pos_insercao)
    }


def aplicar_formulas_dinamicas(ws_base, colunas_meses, base_wb):
    """
    Aplica fórmulas na BASE usando a estratégia de APPEND (Cirúrgica) para L, M e N.
    """
    # 1. Preparação Básica
    ultima_linha = ws_base.max_row
    while ultima_linha > 1 and ws_base.cell(row=ultima_linha, column=1).value is None:
        ultima_linha -= 1

    if ultima_linha < 2 or not colunas_meses:
        return 0

    target_month_sheet = colunas_meses[-1]['nome']
    print(f"DEBUG: Atualizando fórmulas (L, M, N) para incluir aba: {target_month_sheet}")

    # ==============================================================================
    # 🩹 1. PREPARAR TEMPLATE DA COLUNA L (Parcela Paga? - Sim/Não)
    # ==============================================================================
    cell_l2 = ws_base.cell(row=2, column=12)
    formula_l_base = str(cell_l2.value) if cell_l2.value else ""
    formula_l_limpa = formula_l_base.replace(";", ",") # Padroniza para vírgula
    
    nova_formula_l = formula_l_limpa
    
    if not formula_l_limpa.startswith("="):
         # Cria do zero se vazia
         nova_formula_l = f'=IF(OR(NOT(ISERROR(VLOOKUP(A2,\'{target_month_sheet}\'!A:A,1,0)))),"Sim","Não")'
    elif target_month_sheet not in formula_l_limpa:
        # Procura o fechamento: ),"Sim"
        marcador_l = '),"Sim"'
        if marcador_l in formula_l_limpa:
            # Insere: ,NOT(ISERROR(VLOOKUP(A2,'MES'!A:A,1,0)))
            novo_trecho_l = f",NOT(ISERROR(VLOOKUP(A2,'{target_month_sheet}'!A:A,1,0)))"
            nova_formula_l = formula_l_limpa.replace(marcador_l, novo_trecho_l + marcador_l)
            print("✅ Coluna L: Fórmula atualizada.")

    # ==============================================================================
    # 🩹 2. PREPARAR TEMPLATE DA COLUNA M (Data Pagamento - IFERROR)
    # ==============================================================================
    cell_m2 = ws_base.cell(row=2, column=13)
    formula_m_base = str(cell_m2.value) if cell_m2.value else ""
    formula_m_limpa = formula_m_base.replace(";", ",")
    
    nova_formula_m = formula_m_limpa

    if not formula_m_limpa.startswith("="):
        nova_formula_m = '="Pendente de pagamento"'
    
    if target_month_sheet not in formula_m_limpa:
        marcador_m = '"Pendente de pagamento"'
        if marcador_m in formula_m_limpa:
            # Insere: IFERROR(VLOOKUP(A2,'MES'!A:N,14,0), ...
            trecho_novo_m = f"IFERROR(VLOOKUP(A2,'{target_month_sheet}'!A:N,14,0), "
            nova_formula_m = formula_m_limpa.replace(marcador_m, trecho_novo_m + marcador_m) + ")"
            print("✅ Coluna M: Fórmula atualizada.")

    # ==============================================================================
    # 🩹 3. PREPARAR TEMPLATE DA COLUNA N (Soma Countif)
    # ==============================================================================
    cell_n2 = ws_base.cell(row=2, column=14)
    formula_n_base = str(cell_n2.value) if cell_n2.value else ""
    formula_n_limpa = formula_n_base.replace(";", ",") # Vírgula sempre!
    
    nova_formula_n = formula_n_limpa
    
    if not formula_n_limpa.startswith("="):
        # Se vazia, cria a primeira: =COUNTIF('FEV.26'!A:A,BASE!A2)
        nova_formula_n = f"=COUNTIF('{target_month_sheet}'!A:A,BASE!A2)"
    
    elif target_month_sheet not in formula_n_limpa:
        # Se já existe, é só adicionar no final: +COUNTIF(...)
        # Nota: Usamos BASE!A2 como referência template
        novo_trecho_n = f"+COUNTIF('{target_month_sheet}'!A:A,BASE!A2)"
        nova_formula_n += novo_trecho_n
        print("✅ Coluna N: Fórmula atualizada.")

    # ==============================================================================
    # 🚀 APLICAÇÃO EM MASSA (Arrastar Templates)
    # ==============================================================================
    linhas_processadas = 0
    
    for row in range(2, ultima_linha + 1):
        
        # O segredo aqui é o REPLACE inteligente.
        # Estamos usando as fórmulas da linha 2 (A2) como molde.
        # Trocamos "A2" por "A{row}" (ex: A310) para aplicar na linha certa.
        
        # --- COLUNA L ---
        f_l = nova_formula_l.replace("A2", f"A{row}")
        ws_base.cell(row=row, column=12, value=f_l)
        
        # --- COLUNA M ---
        f_m = nova_formula_m.replace("A2", f"A{row}")
        ws_base.cell(row=row, column=13, value=f_m)
        
        # --- COLUNA N ---
        # Substitui BASE!A2 por BASE!A{row}
        # O replace simples funciona bem aqui
        f_n = nova_formula_n.replace("A2", f"A{row}")
        ws_base.cell(row=row, column=14, value=f_n)

        # Copiar Estilo (Visual apenas)
        if row > 2:
            try:
                for col in [12, 13, 14]:
                    copiar_estilo(ws_base.cell(row-1, col), ws_base.cell(row, col))
            except: pass
            
        linhas_processadas += 1
        
    return linhas_processadas

def processar_inadimplentes(dados_filtrados, ws_destino, base_wb, nome_coluna_id):
    """
    Simula o MATCH(V, Q:Q, 0) em Python. 
    Lê a Coluna Q, verifica quais IDs filtrados NÃO estão nela e envia para INADIMPLENTES.
    """

    # --- FUNÇÃO AUXILIAR PARA LIMPAR IDs ---
    def limpar_id(valor):
        if pd.isna(valor): return ""
        return str(valor).strip().replace('.0', '')

    # 1. Pega valores da Coluna Q
    valores_coluna_q = set()
    for celula in ws_destino['Q']:
        if celula.value is not None:
            valores_coluna_q.add(limpar_id(celula.value))
            
    # 2. Pega os IDs do DataFrame
    ids_novos = dados_filtrados[nome_coluna_id].dropna().apply(limpar_id)
    
    # 3. Lógica do "Não": Encontrar Inadimplentes
    ids_inadimplentes = []
    for id_val in ids_novos.unique():
        if id_val and id_val not in valores_coluna_q:
            ids_inadimplentes.append(id_val)
            
    # 4. Grava os dados completos na aba INADIMPLENTES
    if ids_inadimplentes:
        if 'INADIMPLENTES' not in base_wb.sheetnames:
            print("❌ Erro: Aba 'INADIMPLENTES' não encontrada na planilha.")
            raise ValueError("Aba 'INADIMPLENTES' não encontrada na planilha.")
            
        ws_inad = base_wb['INADIMPLENTES']
        ws_base = base_wb['BASE']
        linha_destino = ws_inad.max_row + 1

        # Roubar as fórmulas atualizadas da BASE (linha 2) para usar como molde
        formula_molde_L = str(ws_base.cell(row=2, column=12).value or "")
        formula_molde_M = str(ws_base.cell(row=2, column=13).value or "")

        df_temp = dados_filtrados.copy()
        df_temp['ID_LIMPO'] = df_temp[nome_coluna_id].apply(limpar_id)
        df_inadimplentes = df_temp[df_temp['ID_LIMPO'].isin(ids_inadimplentes)]
        df_inadimplentes = df_inadimplentes.drop_duplicates(subset=['ID_LIMPO'])

        # --- FUNÇÃO AUXILIAR PARA CÁLCULOS ---
        def extrair_numero(val):
            if pd.isna(val) or str(val).strip().startswith('='): return 0.0
            if isinstance(val, (int, float)): return float(val)
            val_str = str(val).replace('.', '').replace(',', '.')
            try: return float(val_str)
            except: return 0.0
        
        for index, row in df_inadimplentes.iterrows():
            valores_linha = row.iloc[0:16].tolist()
            
            # Recálculos das "Fotografias"
            val_valor_emp  = extrair_numero(valores_linha[3]) # D
            val_parcelas   = extrair_numero(valores_linha[4]) # E
            val_fee        = extrair_numero(valores_linha[8]) # I
            val_recebidas  = extrair_numero(valores_linha[13]) # N
            
            valores_linha[7] = valores_linha[5] # H recebe F
            valores_linha[9] = val_valor_emp * val_fee # J
            valores_linha[14] = (val_recebidas / val_parcelas) if val_parcelas > 0 else 0.0 # O
            valores_linha[15] = max(0, val_parcelas - val_recebidas) # P

            for col_idx, valor in enumerate(valores_linha, start=1):
                
                texto_valor = str(valor).strip()

                # =========================================================
                # O CÉREBRO DA FORMATAÇÃO (BLOCO ÚNICO E LIMPO)
                # =========================================================
                
                # 1. Injeção Dinâmica das Fórmulas em L (12) e M (13)
                if col_idx == 12 and formula_molde_L.startswith('='):
                    valor_excel = re.sub(r'\$?[Aa]\$?\d+', f'A{linha_destino}', formula_molde_L)
                    
                elif col_idx == 13 and formula_molde_M.startswith('='):
                    valor_excel = re.sub(r'\$?[Aa]\$?\d+', f'A{linha_destino}', formula_molde_M)
                    
                # 2. Barreira de proteção (Para o resto, se for nulo ou fórmula velha = Branco)
                elif pd.isna(valor) or texto_valor.startswith('='):
                    valor_excel = None

                # 3. Limpeza EXCLUSIVA da Coluna H (8) para tirar horas
                elif col_idx == 8:
                    if isinstance(valor, str) and 'T' in valor:
                        data_str = valor.split('T')[0]
                        try: valor_excel = pd.to_datetime(data_str).date()
                        except: valor_excel = data_str
                    elif isinstance(valor, pd.Timestamp):
                        valor_excel = valor.date()
                    else:
                        valor_excel = valor

                # 4. Tratamento de outras datas padrão do Pandas
                elif isinstance(valor, pd.Timestamp):
                    valor_excel = valor.to_pydatetime()
                
                # 5. Qualquer outro dado passa direto
                else:
                    valor_excel = valor

                # =========================================================

                # Escrever na Célula
                celula_nova = ws_inad.cell(row=linha_destino, column=col_idx, value=valor_excel)
                
                # Clonar estilo
                if linha_destino > 2:
                    celula_referencia = ws_inad.cell(row=linha_destino - 1, column=col_idx)
                    try: copiar_estilo(celula_referencia, celula_nova)
                    except: pass 
                
                # Formatação Numérica Nativa
                if col_idx == 6:  # F (Data com Hora)
                    celula_nova.number_format = 'yyyy-mm-ddThh:mm:ss'
                elif col_idx in [8, 13]:  # H e M (Datas Curtas)
                    celula_nova.number_format = 'dd/mm/yyyy'
                elif col_idx in [4, 10]: # D e J (Moeda/Valores)
                    celula_nova.number_format = '#,##0.00'
                elif col_idx in [9, 15]: # I e O (Porcentagem)
                    celula_nova.number_format = '0.00%'
                    
            linha_destino += 1
            
        print(f"⚠️ {len(ids_inadimplentes)} Inadimplentes encontrados! Dados transferidos com sucesso.")
    else:
        print("✅ Nenhum inadimplente encontrado neste ciclo.")

    if 'INADIMPLENTES' in base_wb.sheetnames:
        ws_inad = base_wb['INADIMPLENTES']
        # Aplica o filtro em toda a extensão da tabela (ex: A1:P500)
        ws_inad.auto_filter.ref = ws_inad.dimensions
        
    return base_wb

def remover_pagantes_inadimplentes(arquivo_base, base_wb):
    """
    Lê a planilha com data_only=True para avaliar as fórmulas.
    Descobre quem pagou ('Sim' na Coluna L) e deleta a linha no base_wb oficial.
    """
    # 1. Carrega uma "cópia fantasma" apenas para ler os valores (resultados)
    arquivo_base.seek(0)
    wb_valores = openpyxl.load_workbook(BytesIO(arquivo_base.read()), data_only=True)
    
    if 'INADIMPLENTES' not in wb_valores.sheetnames:
        return base_wb, 0
        
    ws_valores = wb_valores['INADIMPLENTES']
    ws_oficial = base_wb['INADIMPLENTES']
    
    linhas_para_deletar = []
    
    # 2. Varre DE BAIXO PARA CIMA (Muito importante!)
    for row in range(ws_valores.max_row, 1, -1):
        # Coluna L é a 12
        valor_col_L = ws_valores.cell(row=row, column=12).value
        
        if str(valor_col_L).strip().upper() == 'SIM':
            linhas_para_deletar.append(row)
            
    # 3. Deleta as linhas na planilha oficial que tem as fórmulas
    for row in linhas_para_deletar:
        ws_oficial.delete_rows(row)
        
    return base_wb, len(linhas_para_deletar)


def processar_ciclo_validacao(base_df, base_wb, target_month_name, data_inicio, data_fim):
    """
    Versão FINAL: 
    1. Usa COLUNA FIXA para a Data de Origem (Você define a letra).
    2. Grava nos destinos V e W (Corrigindo o legado T/U).
    3. Cria fórmula em X.
    """
    
    # ==============================================================================
    # ⚙️ CONFIGURAÇÃO: QUAL A LETRA DA COLUNA DE DATA NA BASE?
    # ==============================================================================
    COLUNA_DATA_LETRA = 'F'  # <--- ALTERE AQUI SE FOR 'H' ou 'D'
    # ==============================================================================

    print("#### 🕵️ Diagnóstico do Ciclo")
    
    # 1. Preparar Aba Destino
    if target_month_name not in base_wb.sheetnames:
        print(f"❌ Erro: Aba {target_month_name} não encontrada.")
        raise ValueError(f"Aba {target_month_name} não encontrada.")
    ws_destino = base_wb[target_month_name]
    
    # 2. Localizar as Colunas de Origem (Baseadas na Letra)
    try:
        # Converte letra 'F' para índice numérico (F -> 6). 
        # Pandas usa base 0, então subtraímos 1 (F vira índice 5).
        idx_data = column_index_from_string(COLUNA_DATA_LETRA) - 1
        
        # Pega o nome do cabeçalho dessa coluna para usar no Pandas
        nome_coluna_data = base_df.columns[idx_data]
        nome_coluna_id = base_df.columns[0] # Assume Coluna A (ID/CCB)
        
        print(f"📍 Lendo datas da Coluna **{COLUNA_DATA_LETRA}** (Cabeçalho: '{nome_coluna_data}')")
        
    except IndexError:
        print(f"❌ Erro: A coluna {COLUNA_DATA_LETRA} não existe na planilha BASE.")
        raise ValueError(f"A coluna {COLUNA_DATA_LETRA} não existe na planilha BASE.")

    # 3. Tratamento e Filtro
    try:
        # Força conversão para data (ignora erros de texto/sujeira)
        # dayfirst=True ajuda o Excel brasileiro (DD/MM/AAAA)
        coluna_datas_limpas = pd.to_datetime(base_df[nome_coluna_data], errors='coerce', format='mixed').dt.date
        
        # Mostra exemplo para você conferir
        exemplo = coluna_datas_limpas.dropna().iloc[0] if not coluna_datas_limpas.dropna().empty else "Vazio"
        print(f"🔎 Exemplo de data lida: {exemplo}")
        
        # Aplica o Filtro
        mask = (coluna_datas_limpas >= data_inicio) & (coluna_datas_limpas <= data_fim)
        dados_filtrados = base_df[mask].copy()
        
        qtd = len(dados_filtrados)
        
        if qtd == 0:
            print(f"⚠️ Nenhuma linha encontrada entre {data_inicio} e {data_fim}.")
        else:
            print(f"✅ Filtro OK! {qtd} registros encontrados.")

    except Exception as e:
        print(f"Erro ao processar datas: {e}")
        raise ValueError(f"Erro ao processar datas: {e}")

    # 4. Escrever Dados nos Destinos V (22), W (23) e X (24)
    # Limpa área antiga (garantia para não misturar dados)
    max_row = ws_destino.max_row
    if max_row >= 2:
        # Limpa V, W, X
        for row in ws_destino.iter_rows(min_row=2, max_row=max_row, min_col=22, max_col=24):
            for cell in row:
                cell.value = None
    
    # Loop de Escrita
    linha_atual = 2
    for index, row in dados_filtrados.iterrows():
        
        # --- DESTINO COLUNA V (22) ---
        # Recebe o ID (Coluna A da Base)
        ws_destino.cell(row=linha_atual, column=22).value = row[nome_coluna_id]
        
        # --- DESTINO COLUNA W (23) ---
        # Recebe a Data (Coluna 'F' da Base, limpa)
        val_data = coluna_datas_limpas[index] # Pega a data já tratada
        ws_destino.cell(row=linha_atual, column=23).value = val_data
        
        # --- DESTINO COLUNA X (24) ---
        # Fórmula de Validação
        # Sintaxe OpenPyxl (Inglês + Vírgulas): =IF(ISNUMBER(MATCH(V2,Q:Q,0)),"Sim","Não")
        formula = f'=IF(ISNUMBER(MATCH(V{linha_atual},Q:Q,0)),"Sim","Não")'
        ws_destino.cell(row=linha_atual, column=24).value = formula
        
        # Copiar Estilos (Opcional, pega da linha anterior se existir)
        if linha_atual > 2:
             try:
                 copiar_estilo(ws_destino.cell(linha_atual-1, 22), ws_destino.cell(linha_atual, 22))
                 copiar_estilo(ws_destino.cell(linha_atual-1, 23), ws_destino.cell(linha_atual, 23))
                 copiar_estilo(ws_destino.cell(linha_atual-1, 24), ws_destino.cell(linha_atual, 24))
             except: pass

        linha_atual += 1

    base_wb = processar_inadimplentes(
        dados_filtrados=dados_filtrados, 
        ws_destino=ws_destino, 
        base_wb=base_wb, 
        nome_coluna_id=nome_coluna_id
    )
        
    return qtd


def aplicar_formulas_estaticas(ws_base, linha_inicio):
    """
    Aplica fórmulas estáticas O, P, V nas novas linhas.
    
    Args:
        ws_base: Worksheet da BASE
        linha_inicio: Primeira linha onde começaram os novos dados
    
    Returns:
        int: Número de linhas processadas
    """
    ultima_linha = encontrar_ultima_linha(ws_base)
    
    # Encontrar índice da coluna DATA dinamicamente
    col_data_index = encontrar_coluna_por_header(ws_base, 'DATA')
    
    if not col_data_index:
        raise ValueError(
            "CRÍTICO: Coluna 'DATA' não encontrada na aba BASE. "
            "Verifique se o header da coluna está exatamente como 'DATA' (case-sensitive)."
        )
    
    # Log da coluna encontrada
    col_data_letra = get_column_letter(col_data_index)
    print(f"DEBUG: Coluna 'DATA' encontrada no índice {col_data_index} (letra {col_data_letra})")
    
    linhas_processadas = 0
    
    for row in range(linha_inicio, ultima_linha + 1):
        # Linha molde: linha anterior (row - 1)
        linha_molde = row - 1
        
        # Col O (15) - % Recebimento: =N2/E2
        cell_o = ws_base.cell(row=row, column=15, value=f"=N{row}/E{row}")
        if linha_molde >= 2:
            copiar_estilo(ws_base.cell(row=linha_molde, column=15), cell_o)
        
        # Col P (16) - Pendentes: =E2-N2
        cell_p = ws_base.cell(row=row, column=16, value=f"=E{row}-N{row}")
        if linha_molde >= 2:
            copiar_estilo(ws_base.cell(row=linha_molde, column=16), cell_p)
        
        # Col DATA (índice dinâmico) - Fórmula TEXT para serial number
        # CRÍTICO: Coluna F contém serial number do Excel (ex: 45992.2548)
        # Converter para formato dd/mm/aaaa usando TEXT
        cell_data = ws_base.cell(row=row, column=col_data_index, value=f'=TEXT(F{row},"dd/mm/aaaa")')
        if linha_molde >= 2:
            copiar_estilo(ws_base.cell(row=linha_molde, column=col_data_index), cell_data)
        
        linhas_processadas += 1
    
    return linhas_processadas


def atualizar_aba_base(base_wb, parceiro_wb, target_month, linha_inicio_append):
    """
    Atualiza a aba BASE com novos dados e fórmulas dinâmicas.
    
    IMPORTANTE: As fórmulas dinâmicas (L, M, N) são aplicadas em TODAS as linhas,
    não apenas nas novas, pois registros antigos podem ter pago no novo mês.
    
    Args:
        base_wb: Workbook do arquivo BASE
        parceiro_wb: Workbook do arquivo PARCEIRO
        target_month: String do mês (ex: 'JAN.26')
        linha_inicio_append: Primeira linha onde foram adicionados dados de Produção
                           (usado apenas para fórmulas estáticas O, P, V)
    
    Returns:
        dict: {
            'linhas_producao': int,
            'coluna_mes_inserida': str,
            'abas_meses_encontradas': list,
            'linhas_formulas_aplicadas': int,    # Total de linhas (L, M, N)
            'linhas_novas_estaticas': int        # Apenas novas (O, P, V)
        }
    """
    # 1. Obter referências
    ws_base = base_wb['BASE']
    ws_producao = parceiro_wb['Produção']
    
    # 2. Identificar colunas de meses existentes (entre P e V)
    colunas_meses = encontrar_colunas_meses(ws_base)
    
    # 3. Inserir nova coluna de mês
    col_inserida = inserir_coluna_mes(ws_base, target_month, colunas_meses)
    
    # 4. Atualizar colunas_meses com a nova coluna
    colunas_meses.append(col_inserida)
    
    # 5. Aplicar fórmulas dinâmicas (L, M, N) em TODAS as linhas
    # CRÍTICO: Atualiza todas as linhas, não apenas novas, pois registros
    # antigos podem ter pago no novo mês e precisam ser atualizados
    # CORREÇÃO: Passar base_wb para validação de abas locais
    linhas_processadas = aplicar_formulas_dinamicas(
        ws_base, 
        colunas_meses,
        base_wb  # NOVO: passar workbook para validação
    )
    
    # 6. Aplicar fórmulas estáticas (O, P, V) nas novas linhas
    linhas_novas = aplicar_formulas_estaticas(ws_base, linha_inicio_append)
    
    # 7. Retornar métricas
    return {
        'coluna_mes_inserida': target_month,
        'abas_meses_encontradas': [col['nome'] for col in colunas_meses],
        'linhas_formulas_aplicadas': linhas_processadas,  # L, M, N (todas)
        'linhas_novas_estaticas': linhas_novas           # O, P, V (apenas novas)
    }


def inserir_dados_colunas_especificas(ws_origem, ws_destino, col_inicio=1, col_fim=13, linha_destino_inicio=2):
    """
    Copia dados de worksheet origem para destino, mas apenas em colunas específicas.
    
    Args:
        ws_origem: Worksheet de origem
        ws_destino: Worksheet de destino
        col_inicio: Primeira coluna a copiar (default: 1 = A)
        col_fim: Última coluna a copiar (default: 13 = M)
        linha_destino_inicio: Linha inicial no destino (default: 2)
    
    Returns:
        int: Número de linhas copiadas
    
    Nota:
        Colunas N-X são preenchidas pela função aplicar_regras_colunas_n_x()
    """
    linhas_copiadas = 0
    linha_destino = linha_destino_inicio
    
    # Iterar sobre linhas da origem (pulando header - linha 1)
    for row in ws_origem.iter_rows(min_row=2, values_only=True):
        # Pular linhas vazias
        if all(cell is None for cell in row):
            continue
        
        # Copiar apenas colunas especificadas
        for col_idx in range(col_inicio, min(col_fim + 1, len(row) + 1)):
            valor = row[col_idx - 1] if col_idx <= len(row) else None
            ws_destino.cell(row=linha_destino, column=col_idx, value=valor)
        
        linha_destino += 1
        linhas_copiadas += 1
    
    return linhas_copiadas


def reaplicar_formulas(ws, formulas_dict, linha_inicio=2, linha_fim=None):
    """
    Aplica fórmulas capturadas em um range de linhas, atualizando referências.
    
    Args:
        ws: Worksheet do openpyxl
        formulas_dict: Dict {col_idx: formula_template}
        linha_inicio: Primeira linha onde aplicar (default: 2)
        linha_fim: Última linha (default: None = até última linha com dados)
    
    Returns:
        int: Número de fórmulas aplicadas
    """
    if linha_fim is None:
        linha_fim = encontrar_ultima_linha(ws)
    
    formulas_aplicadas = 0
    
    for linha in range(linha_inicio, linha_fim + 1):
        for col_idx, formula_template in formulas_dict.items():
            # Atualizar referências de linha na fórmula
            formula_atualizada = atualizar_formula_linha(formula_template, linha)
            
            # Aplicar fórmula na célula
            ws.cell(row=linha, column=col_idx, value=formula_atualizada)
            formulas_aplicadas += 1
    
    return formulas_aplicadas

